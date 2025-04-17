import os
import time
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from datetime import datetime

def check_file_online(filepath):
    while True:
        if os.path.exists(filepath) and os.path.isfile(filepath):
            print(f"File {filepath} is online.")
            break
        print("File not online. Retrying...")
        time.sleep(5)

def open_teams_and_link():
    driver = webdriver.Chrome()  # Ensure you have the ChromeDriver installed
    driver.get("https://teams.microsoft.com")
    time.sleep(10)  # Wait for Teams to load
    driver.get("https://maestro.vivo.com.br/sites_fibrados/exports/RELATORIO_BASE_PORTAL_PROJETO_SITES_FIBRA.xlsx")
    time.sleep(10)  # Wait for download to complete
    driver.quit()

def copy_report_to_excel(report_path, target_excel, target_sheet):
    wb_report = load_workbook(report_path)
    ws_report = wb_report.active

    wb_target = load_workbook(target_excel)
    ws_target = wb_target[target_sheet]

    for row in ws_report.iter_rows(values_only=True):
        ws_target.append(row)

    wb_target.save(target_excel)
    print("Report copied successfully.")

def apply_filters_and_fill_data(excel_path, sheet_name):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    today = datetime.now().strftime("%d/%m/%Y")

    # Example: Iterate through column Q and apply filters
    for row in ws.iter_rows(min_row=2, max_col=17, values_only=False):  # Assuming column Q is 17th
        portal_status = row[16].value  # Column Q
        if portal_status in ["Ativos Não Migrados", "Migrados"]:
            entrega_fibra_plan = row[10].value  # Example column
            entrega_fibra_real = row[11].value  # Example column
            if not entrega_fibra_plan or not entrega_fibra_real:
                row[10].value = today
                row[11].value = today

    wb.save(excel_path)
    print("Filters applied and data filled successfully.")

# Main execution
if __name__ == "__main__":
    # Step 1: Check if the file is online
    excel_path = r"C:\Users\A0164663\Telefonica\Mirthes Maria Spinola Bastos - RF x TX\RF_x_TX_BA-SE-NE_ABR_MAI_JUN - MAIO.V3.xlsx"
    check_file_online(excel_path)

    # Step 2: Open Teams and download the report
    open_teams_and_link()

    # Step 3: Copy report content to the target Excel
    report_path = r"C:\Users\A0164663\Downloads\RELATORIO_BASE_PORTAL_PROJETO_SITES_FIBRA.xlsx"
    copy_report_to_excel(report_path, excel_path, "PORTAL FIBRA")

    # Step 4: Apply filters and fill data
    apply_filters_and_fill_data(excel_path, "METRO")
